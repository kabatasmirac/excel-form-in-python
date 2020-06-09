# -*- coding: utf-8 -*-
"""
Created on Tue Jun  9 14:34:48 2020

@author: mirac.kabatas
"""
import os
import tkinter as tk
from tkinter import ttk
from openpyxl import Workbook
import openpyxl

def __init__(self):
    ''' Initialize the main Frame '''
    controls = ttk.Frame(root1,width=700, height = 700+300+10)
    ttk.Label(controls, text='---Doku Açıklamaları---',font=('arial 18')).grid(row=25,column=0,columnspan=2,sticky="W",padx=5,pady=5)
    #controls adında bir frame oluşturup genişlik ve yükseklik verildi.
    #controls framenin içine Label adına bir alan açtık. Form nesnelerimizi Label alanının içine yerleştireceğiz.
    #Label alanının içinde text, font gibi ayarları yapıp grid(ızgaralara) böldük ve padding ayarlarını verdik.
    #sticky="W" -> pusula yönlerine göre controls framesi içindeki Label alanının yerini belirledik.    

    
    controls.grid(row=0,column=4,sticky="E",padx=10)
    #Formunun oluşturulması
    #Form Labelların oluşturulması
    
    attr1=ttk.Label(controls,text="Meslek")
    attr2=ttk.Label(controls,text="Ad Soyad")
    attr3=ttk.Label(controls,text="Çalışma Durumu")
    attr4=ttk.Label(controls,text="Medeni Hali")
    attr5=ttk.Label(controls,text="Çocuk Sayısı")
    attr6=ttk.Label(controls,text="Ehliyet")
    attr7=ttk.Label(controls,text="Kariyer Hedefi")

    #/************************************************************************  
    #form entitylerin oluşturulması
    
    
    #/************************************************************************  
    #optionslist(dropdown selection)
    ent_attr1=tk.StringVar(controls)
    optionslist1=["Mühendis ", "Yüksek Mühendis"]
    option1 = ttk.OptionMenu(controls, ent_attr1,optionslist1[0] ,*optionslist1)#optionslist1[0] ,*optionslist1 yazılmazsa optionlist1[0] elemanını yutuyor
    option1.grid(row=0,column=1, padx=5, pady=5)
    
    
    #/************************************************************************  
    ent_attr2=ttk.Entry(controls)#(Textbox)
    
    
   #/************************************************************************  
   #/*radiobutton
    ent_attr3 = tk.IntVar(controls)
    #tk.IntVar->radiobutton
    R1 = ttk.Radiobutton(controls, text = "Çalışıyor", 
                     variable = ent_attr3, value = 1)
    R2 = ttk.Radiobutton(controls, text = "Çalışmıyor", 
                     variable = ent_attr3, value = 0)
    
    #variable parametresi ile radio butonları gruplayabiliyoruz. Biri seçiliyken diğeri seçilemiyor.
    #radiobuttonname.grid -> radio buttonların konumlarını grid ızgara içerisine yerleştirmeye yarıyor
    R1.grid(row = 3, column = 0, padx = 10)
    R2.grid(row = 3, column = 1, padx = 10)
    #/************************************************************************ 
     #/*radiobutton
    ent_attr4 = tk.IntVar(controls)

    R3 = ttk.Radiobutton(controls, text = "Evli", 
                     variable = ent_attr4, value = 1)
    R4 = ttk.Radiobutton(controls, text = "Bekar", 
                     variable = ent_attr4, value = 0)
    R3.grid(row = 5, column = 0, padx = 10)
    R4.grid(row = 5, column = 1, padx = 10)
    #/************************************************************************ 
     #/*radiobutton
    ent_attr5 = tk.IntVar(controls)

    R5 = ttk.Radiobutton(controls, text = "0", 
                     variable = ent_attr5, value = 1)
    R6 = ttk.Radiobutton(controls, text = "0+", 
                     variable = ent_attr5, value = 0)
    
    
    R5.grid(row = 7, column = 0, padx = 10)
    R6.grid(row = 7, column = 1, padx = 10)
    #/************************************************************************  
     #/*radiobutton
    ent_attr6 = tk.IntVar(controls)

    R7 = ttk.Radiobutton(controls, text = "Yok", 
                     variable = ent_attr6, value = 1)
    R8 = ttk.Radiobutton(controls, text = "Var", 
                     variable = ent_attr6, value = 0)
    
    
    R7.grid(row = 9, column = 0, padx = 10)
    R8.grid(row = 9, column = 1, padx = 10)
    
    #/************************************************************************  
    #textboxların oluşturulması
    ent_attr7=ttk.Entry(controls)
    
    #attr.grid(row=0,column=0, padx=5, pady=5)
    #grid : form için ayrılan alanı gridlere (ızgaralara) bölüyor
    #row,column : hangi satır ve sutuna yerleştirmek istiyorsak onu belirlediğimiz parametreler
    #padx,pady : padding parametreleri
    
    attr1.grid(row=0,column=0, padx=5, pady=5)
    attr2.grid(row=1,column=0, padx=5, pady=5)
    ent_attr2.grid(row=1,column=1, padx=5, pady=5)
    
    attr3.grid(row=2,column=0,columnspan=3, padx=5, pady=5)
    attr4.grid(row=4,column=0,columnspan=3, padx=5, pady=5)
    attr5.grid(row=6,column=0,columnspan=3, padx=5, pady=5)
    attr6.grid(row=8,column=0,columnspan=3, padx=5, pady=5)
    attr7.grid(row=10,column=0, padx=5, pady=5)
    ent_attr7.grid(row=10,column=1, padx=5, pady=5)
    
    
    attrbtn=ttk.Button(controls,text="Save!",command=addsheet)
    attrbtn.grid(row=18,column=1,columnspan=2,padx=5,pady=5)
    
    
    ttk.Frame.__init__(self,)
    
    #attr1.grid(row=0,column=0, padx=5, pady=5)
    #grid : form için ayrılan alanı gridlere (ızgaralara) bölüyor
    #row,column : hangi satır ve sutuna yerleştirmek istiyorsak onu belirlediğimiz parametreler
    #padx,pady : padding parametreleri
    
def addsheet():
    sheetname="tablo.xlsx"
    
    if os.path.exists(sheetname) == False:
        #dosya yoksa oluşturulur
        book = Workbook()
        sheet = book.active
        sheet["A1"]='Meslek'
        sheet["B1"]='Ad_Soyad'
        sheet["C1"]='Calisma_Durumu'
        sheet["D1"]='Medeni_Hali'
        sheet["E1"]='Cocuk_Sayisi'
        sheet["F1"]='Ehliyet'
        sheet["G1"]='Kariyer_Hedefi'
        #dosya oluşturulurken sayfa başlığını şekildeki gibi ekleyip bir alttaki kod satırıyla kaydediyoruz.
        book.save(sheetname)
        print("Dosya oluşturuldu..\nFile created..")
    
    sheet1 = openpyxl.load_workbook(filename = sheetname)
    sheet=sheet1.active
    #yukarıdaki satırlarda kayıt edildilten sonra kapanan excel sayfamızı yeniden açıp yazmak için aktif hale getirdik.
    
    name1=ent_attr1.get()
    name2=ent_attr2.get()
    name3=ent_attr3.get()
    name4=ent_attr4.get()
    name5=ent_attr5.get()
    name6=ent_attr6.get()
    name7=ent_attr7.get()
    #yukarıdaki kod satırlarında entity içerisinden alınan bilgileri değişkenler içinde tutuyoruz.
    
    max_rowi=sheet.max_row
    #datasheet içindeki satır sayıları kaç adet kayıt var onun sayısını döndürür
    #+1 demek ise bizim kaydımızı yazmak istediğimiz satır
    
    sheet["A"+str(max_rowi+1)]=name1
    sheet["B"+str(max_rowi+1)]=name2
    sheet["C"+str(max_rowi+1)]=name3
    sheet["D"+str(max_rowi+1)]=name4
    sheet["E"+str(max_rowi+1)]=name5
    sheet["F"+str(max_rowi+1)]=name6
    sheet["G"+str(max_rowi+1)]=name7
    #ilgili hücrenin herbirine gelecek olan değişkenler içinde tuttuğumuz kullanıcıdan alınan bilgileri yazdık
    
    
    sheet1.save(sheetname)
    #save diyerek kaydedip kapattık
    print("Form gönderildi..")

      
from tkinter import *        
if __name__ == '__main__':
    
    root1 = Tk()
    app1 = Window(root1)
    root1.geometry("700x700+300+10")
    root1.wm_title("Tool")
    root1.mainloop()